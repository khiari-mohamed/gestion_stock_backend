from typing import List
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExportService:
    """Service pour générer des exports PDF et Excel avec formatage professionnel"""
    
    def _format_worksheet(self, worksheet, has_total_row: bool = False):
        """Appliquer un formatage professionnel à une feuille Excel"""
        # Style pour l'en-tête
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # Style pour la ligne de total
        total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        total_font = Font(bold=True, size=11)
        
        # Bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Formater l'en-tête
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Formater les données
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Formater les nombres
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.000'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Formater la ligne de total si présente
        if has_total_row:
            for cell in worksheet[worksheet.max_row]:
                cell.fill = total_fill
                cell.font = total_font
                cell.border = thin_border
        
        # Ajuster la largeur des colonnes
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 3, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Figer la première ligne
        worksheet.freeze_panes = 'A2'
    
    def export_mouvements_excel(self, mouvements: List) -> bytes:
        """Exporter les mouvements en Excel avec formatage professionnel"""
        data = []
        for m in mouvements:
            data.append({
                "Date": m.date_mouvement.strftime("%d/%m/%Y %H:%M"),
                "Type": m.type,
                "Article": m.article.designation if hasattr(m, 'article') else "",
                "Code": m.article.code if hasattr(m, 'article') else "",
                "Quantité": m.quantite,
                "Prix Unitaire (DT)": round(m.prix_unitaire or 0, 3),
                "Valeur Totale (DT)": round(m.valeur_totale or 0, 3),
                "Motif": m.motif or ""
            })
        
        df = pd.DataFrame(data)
        buffer = BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Mouvements')
            worksheet = writer.sheets['Mouvements']
            self._format_worksheet(worksheet)
        
        return buffer.getvalue()
    
    def export_stock_excel(self, articles: List, magasin_nom: str) -> bytes:
        """Exporter l'état du stock en Excel avec formatage professionnel"""
        data = []
        total_valeur = 0
        
        for a in articles:
            valeur = a.stock_actuel * a.prix_achat
            total_valeur += valeur
            
            # Statut du stock
            if a.stock_actuel == 0:
                statut = "RUPTURE"
            elif a.stock_actuel <= a.stock_min:
                statut = "FAIBLE"
            elif a.stock_actuel >= a.stock_max:
                statut = "EXCESSIF"
            else:
                statut = "NORMAL"
            
            data.append({
                "Code": a.code,
                "Désignation": a.designation,
                "Catégorie": a.categorie or "",
                "Stock Actuel": a.stock_actuel,
                "Stock Min": a.stock_min,
                "Stock Max": a.stock_max,
                "Statut": statut,
                "Prix Achat HT (DT)": round(a.prix_achat, 3),
                "Prix Vente HT (DT)": round(a.prix_vente, 3),
                "Valeur Stock (DT)": round(valeur, 3)
            })
        
        df = pd.DataFrame(data)
        
        # Ajouter ligne de total
        total_row = pd.DataFrame([{
            "Code": "TOTAL",
            "Désignation": "",
            "Catégorie": "",
            "Stock Actuel": sum(a.stock_actuel for a in articles),
            "Stock Min": "",
            "Stock Max": "",
            "Statut": "",
            "Prix Achat HT (DT)": "",
            "Prix Vente HT (DT)": "",
            "Valeur Stock (DT)": round(total_valeur, 3)
        }])
        
        df = pd.concat([df, total_row], ignore_index=True)
        
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Stock')
            worksheet = writer.sheets['Stock']
            self._format_worksheet(worksheet, has_total_row=True)
            
            # Ajouter titre
            worksheet.insert_rows(1)
            worksheet['A1'] = f'État du Stock - {magasin_nom} - {datetime.now().strftime("%d/%m/%Y")}'
            worksheet['A1'].font = Font(bold=True, size=14)
            worksheet.merge_cells('A1:J1')
        
        return buffer.getvalue()
    
    def export_stock_pdf(self, articles: List, magasin_id: str) -> bytes:
        """Exporter l'état du stock en PDF - Utilise Excel pour MVP"""
        # Pour MVP, retourner Excel. Phase 2: implémenter avec ReportLab
        return self.export_stock_excel(articles, f"Magasin {magasin_id}")
    
    def export_valorisation_excel(self, articles: List) -> bytes:
        """Exporter la valorisation du stock (format comptabilité tunisienne)"""
        data = []
        total_valeur_ht = 0
        total_tva = 0
        
        for a in articles:
            valeur_ht = a.stock_actuel * a.prix_achat
            tva = valeur_ht * (a.tva_taux or 0.19)
            valeur_ttc = valeur_ht + tva
            
            total_valeur_ht += valeur_ht
            total_tva += tva
            
            data.append({
                "Code Article": a.code,
                "Désignation": a.designation,
                "Quantité en Stock": a.stock_actuel,
                "Prix Achat HT (DT)": round(a.prix_achat, 3),
                "Valeur Stock HT (DT)": round(valeur_ht, 3),
                "Taux TVA": f"{(a.tva_taux or 0.19) * 100:.0f}%",
                "TVA (DT)": round(tva, 3),
                "Valeur TTC (DT)": round(valeur_ttc, 3)
            })
        
        df = pd.DataFrame(data)
        
        # Ajouter ligne de total
        total_ttc = total_valeur_ht + total_tva
        total_row = pd.DataFrame([{
            "Code Article": "TOTAL",
            "Désignation": "",
            "Quantité en Stock": sum(a.stock_actuel for a in articles),
            "Prix Achat HT (DT)": "",
            "Valeur Stock HT (DT)": round(total_valeur_ht, 3),
            "Taux TVA": "",
            "TVA (DT)": round(total_tva, 3),
            "Valeur TTC (DT)": round(total_ttc, 3)
        }])
        
        df = pd.concat([df, total_row], ignore_index=True)
        
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Valorisation Stock')
            worksheet = writer.sheets['Valorisation Stock']
            self._format_worksheet(worksheet, has_total_row=True)
            
            # Ajouter titre et date
            worksheet.insert_rows(1)
            worksheet.insert_rows(1)
            worksheet['A1'] = 'VALORISATION DU STOCK'
            worksheet['A1'].font = Font(bold=True, size=16, color="366092")
            worksheet['A2'] = f'Date d\'extraction: {datetime.now().strftime("%d/%m/%Y à %H:%M")}'
            worksheet['A2'].font = Font(italic=True, size=10)
            worksheet.merge_cells('A1:H1')
            worksheet.merge_cells('A2:H2')
        
        return buffer.getvalue()
